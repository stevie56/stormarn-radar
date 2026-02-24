"""
app.py – Streamlit Dashboard für den Stormarn KI-Radar
Starte mit: streamlit run app.py
"""
import json
import os
import streamlit as st
import pandas as pd
import folium
from streamlit_folium import st_folium

import config_loader as cfg
import database as db
import scraper
import analyzer
import geo_mapper
import alert
import pdf_export
import bulk_analyzer
import company_finder
import wirtschaftsdaten_importer
import ki_scorer
import news_monitor
import regional_compare
import report_generator
import job_radar
import reanalyzer
import weekly_report

# ──────────────────────────────────────────────────────────
# Seitenkonfiguration
# ──────────────────────────────────────────────────────────
st.set_page_config(
    page_title=cfg.get("radar.name", "Regional Radar"),
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Datenbank initialisieren
db.init_db()

# ──────────────────────────────────────────────────────────
# Styling
# ──────────────────────────────────────────────────────────
PRIMARY = cfg.get("radar.pdf.primary_color", "#1a5276")
ACCENT = cfg.get("radar.pdf.accent_color", "#2e86c1")

st.markdown(f"""
<style>
    .main-header {{
        background: linear-gradient(135deg, {PRIMARY} 0%, {ACCENT} 100%);
        padding: 1.5rem 2rem;
        border-radius: 12px;
        color: white;
        margin-bottom: 1.5rem;
    }}
    .metric-card {{
        background: white;
        border: 1px solid #e0e0e0;
        border-radius: 10px;
        padding: 1rem;
        text-align: center;
        box-shadow: 0 2px 8px rgba(0,0,0,0.06);
    }}
    .badge {{
        display: inline-block;
        padding: 3px 10px;
        border-radius: 12px;
        font-size: 0.8rem;
        font-weight: bold;
        color: white;
    }}
    .badge-echt {{ background: #1e8449; }}
    .badge-int {{ background: #d4ac0d; }}
    .badge-buzz {{ background: #e67e22; }}
    .badge-kein {{ background: #95a5a6; }}
    .stButton>button {{
        background-color: {PRIMARY};
        color: white;
        border: none;
        border-radius: 8px;
    }}
</style>
""", unsafe_allow_html=True)

# ──────────────────────────────────────────────────────────
# Header
# ──────────────────────────────────────────────────────────
radar_name = cfg.get("radar.name", "Regional Radar")
region = cfg.get("radar.region", "")
topic = cfg.get("radar.topic", "")

st.markdown(f"""
<div class="main-header">
    <h1 style="margin:0;font-size:2rem;">🎯 {radar_name}</h1>
    <p style="margin:0.3rem 0 0 0;opacity:0.85;">{region} · {topic} · Live-Dashboard</p>
</div>
""", unsafe_allow_html=True)

# ──────────────────────────────────────────────────────────
# Sidebar – Navigation
# ──────────────────────────────────────────────────────────
with st.sidebar:
    st.image("https://via.placeholder.com/250x60/1a5276/ffffff?text=Stormarn+Radar",
             use_column_width=True)
    st.markdown("---")
    page = st.radio(
        "Navigation",
        ["📊 Dashboard", "🏢 Unternehmen", "🗺️ Karte",
         "➕ Neu analysieren", "📤 Excel-Import", "📊 Wirtschaftsdaten",
         "🔍 Auto-Suche", "💼 Job-Radar", "📰 News-Monitor",
         "🏅 KI-Ranking", "📈 Trends", "🌍 Regionalvergleich",
         "🔬 Qualitäts-Check", "📋 Aktivitätslog", "📄 PDF-Export", "📑 IHK-Bericht",
         "📧 Wochenbericht", "⚙️ Einstellungen"],
        label_visibility="collapsed"
    )
    st.markdown("---")
    api_key = st.text_input("OpenAI API Key", type="password",
                             value=os.getenv("OPENAI_API_KEY", ""),
                             help="Wird nur für die Analyse benötigt")
    if api_key:
        os.environ["OPENAI_API_KEY"] = api_key

# ──────────────────────────────────────────────────────────
# Hilfsfunktionen
# ──────────────────────────────────────────────────────────
CATEGORY_COLORS_HEX = {
    "ECHTER_EINSATZ": "#1e8449",
    "INTEGRATION":    "#d4ac0d",
    "BUZZWORD":       "#e67e22",
    "KEIN_KI":        "#95a5a6",
    "UNBEKANNT":      "#bdc3c7",
}
CATEGORY_LABELS = {
    "ECHTER_EINSATZ": "✅ Echter Einsatz",
    "INTEGRATION":    "🔗 Integration",
    "BUZZWORD":       "⚠️ Buzzword",
    "KEIN_KI":        "❌ Kein KI-Bezug",
    "UNBEKANNT":      "❓ Unbekannt",
}


def category_badge(kat):
    color = CATEGORY_COLORS_HEX.get(kat, "#ccc")
    label = CATEGORY_LABELS.get(kat, kat)
    return f'<span style="background:{color};padding:3px 10px;border-radius:12px;color:white;font-size:0.8rem;font-weight:bold;">{label}</span>'


# ──────────────────────────────────────────────────────────
# PAGE: Dashboard
# ──────────────────────────────────────────────────────────
if page == "📊 Dashboard":
    stats = db.get_stats()
    companies = db.get_all_companies()
    events = db.get_recent_events(5)

    # KPI-Reihe
    col1, col2, col3, col4, col5 = st.columns(5)
    total = stats["total_companies"]
    by_cat = stats["by_category"]

    with col1:
        st.metric("Unternehmen gesamt", total)
    with col2:
        st.metric("✅ Echter Einsatz", by_cat.get("ECHTER_EINSATZ", 0))
    with col3:
        st.metric("🔗 Integration", by_cat.get("INTEGRATION", 0))
    with col4:
        st.metric("⚠️ Buzzword", by_cat.get("BUZZWORD", 0))
    with col5:
        st.metric("❌ Kein KI", by_cat.get("KEIN_KI", 0))

    st.markdown("---")
    col_left, col_right = st.columns([2, 1])

    with col_left:
        st.subheader("🏢 Zuletzt analysierte Unternehmen")
        if companies:
            recent = sorted(companies,
                            key=lambda x: x.get("updated_at", ""),
                            reverse=True)[:8]
            for c in recent:
                with st.container():
                    c1, c2, c3 = st.columns([3, 2, 1])
                    with c1:
                        st.write(f"**{c['name']}**")
                        st.caption(f"{c.get('city', '')} · {c.get('industry', '')}")
                    with c2:
                        kat = c.get("kategorie", "UNBEKANNT")
                        st.markdown(category_badge(kat), unsafe_allow_html=True)
                    with c3:
                        st.write(f"Score: {c.get('vertrauen', '–')}")
        else:
            st.info("Noch keine Unternehmen analysiert. Gehe zu **➕ Neu analysieren**.")

    with col_right:
        st.subheader("📋 Letzte Aktivitäten")
        if events:
            for e in events:
                icon = "🆕" if e["event_type"] == "NEU" else "🔄"
                st.markdown(f"{icon} **{e.get('company_name', '')}**")
                st.caption(f"{e['message']} · {e['created_at'][:16]}")
                st.markdown("---")
        else:
            st.info("Keine Aktivitäten")

# ──────────────────────────────────────────────────────────
# PAGE: Unternehmen
# ──────────────────────────────────────────────────────────
elif page == "🏢 Unternehmen":
    st.header("Unternehmensverzeichnis")
    companies = db.get_all_companies()

    if not companies:
        st.info("Noch keine Unternehmen. Füge sie unter **➕ Neu analysieren** hinzu.")
        st.stop()

    # Filter
    col_f1, col_f2, col_f3 = st.columns(3)
    with col_f1:
        search = st.text_input("🔍 Suche", placeholder="Firmenname...")
    with col_f2:
        all_cats = ["Alle"] + list(CATEGORY_LABELS.values())
        cat_filter = st.selectbox("Kategorie", all_cats)
    with col_f3:
        industries = ["Alle"] + list({c.get("industry", "") for c in companies if c.get("industry")})
        ind_filter = st.selectbox("Branche", industries)

    # Filtern
    filtered = companies
    if search:
        filtered = [c for c in filtered if search.lower() in c["name"].lower()]
    if cat_filter != "Alle":
        kat_key = {v: k for k, v in CATEGORY_LABELS.items()}.get(cat_filter)
        if kat_key:
            filtered = [c for c in filtered if c.get("kategorie") == kat_key]
    if ind_filter != "Alle":
        filtered = [c for c in filtered if c.get("industry") == ind_filter]

    st.markdown(f"**{len(filtered)}** Unternehmen gefunden")
    st.markdown("---")

    for c in filtered:
        with st.expander(f"🏢 {c['name']} · {CATEGORY_LABELS.get(c.get('kategorie',''), '?')}"):
            col1, col2 = st.columns([3, 1])
            with col1:
                if c.get("biografie"):
                    st.markdown(f"*{c['biografie']}*")
                st.markdown(f"**Website:** [{c.get('website','')}]({c.get('website','')})")
                st.markdown(f"**Adresse:** {c.get('address','')}, {c.get('city','')}")
                st.markdown(f"**Branche:** {c.get('industry', '–')}")

                ki_apps = c.get("ki_anwendungen", [])
                if isinstance(ki_apps, str):
                    try:
                        ki_apps = json.loads(ki_apps)
                    except Exception:
                        ki_apps = []
                if ki_apps:
                    st.markdown("**KI-Anwendungen:** " + " · ".join([f"`{a}`" for a in ki_apps]))
            with col2:
                kat = c.get("kategorie", "UNBEKANNT")
                st.markdown(category_badge(kat), unsafe_allow_html=True)
                st.metric("Vertrauen", f"{c.get('vertrauen', '–')}/100")
                if c.get("lat") and c.get("lng"):
                    st.caption(f"📍 {c['lat']:.4f}, {c['lng']:.4f}")

# ──────────────────────────────────────────────────────────
# PAGE: Karte
# ──────────────────────────────────────────────────────────
elif page == "🗺️ Karte":
    st.header("🗺️ KI-Akteure in der Region")
    companies = db.get_all_companies()

    geo_companies = [c for c in companies if c.get("lat") and c.get("lng")]

    if not geo_companies:
        st.warning("Keine Unternehmen mit Koordinaten. Füge Adressen hinzu und analysiere erneut.")
        st.stop()

    # Karte zentrieren
    bounds = cfg.get("radar.region_bounds", {})
    center_lat = (bounds.get("north", 53.7) + bounds.get("south", 53.6)) / 2
    center_lng = (bounds.get("east", 10.25) + bounds.get("west", 10.1)) / 2

    m = folium.Map(location=[center_lat, center_lng], zoom_start=11,
                   tiles="CartoDB positron")

    # Marker-Farbe nach Kategorie
    COLOR_MAP = {
        "ECHTER_EINSATZ": "green",
        "INTEGRATION":    "orange",
        "BUZZWORD":       "red",
        "KEIN_KI":        "gray",
        "UNBEKANNT":      "lightgray",
    }

    for c in geo_companies:
        kat = c.get("kategorie", "UNBEKANNT")
        color = COLOR_MAP.get(kat, "blue")
        ki_apps = c.get("ki_anwendungen", [])
        if isinstance(ki_apps, str):
            try:
                ki_apps = json.loads(ki_apps)
            except Exception:
                ki_apps = []

        popup_html = f"""
        <b>{c['name']}</b><br>
        {CATEGORY_LABELS.get(kat, kat)}<br>
        {c.get('city', '')}<br>
        {'<br>'.join(ki_apps[:3]) if ki_apps else ''}
        <br><a href="{c.get('website','')}" target="_blank">Website öffnen</a>
        """

        folium.CircleMarker(
            location=[c["lat"], c["lng"]],
            radius=10,
            color=color,
            fill=True,
            fill_color=color,
            fill_opacity=0.7,
            popup=folium.Popup(popup_html, max_width=250),
            tooltip=c["name"]
        ).add_to(m)

    # Legende
    legend_html = """
    <div style="position: fixed; bottom: 30px; left: 30px; z-index: 1000;
                background: white; padding: 15px; border-radius: 8px;
                box-shadow: 0 2px 10px rgba(0,0,0,0.2); font-size: 13px;">
        <b>Legende</b><br>
        🟢 Echter KI-Einsatz<br>
        🟠 KI-Integration<br>
        🔴 KI-Buzzword<br>
        ⚫ Kein KI-Bezug
    </div>
    """
    m.get_root().html.add_child(folium.Element(legend_html))

    st_folium(m, width=None, height=550, returned_objects=[])

    col1, col2 = st.columns(2)
    with col1:
        st.metric("Unternehmen auf der Karte", len(geo_companies))
    with col2:
        st.metric("Davon mit echtem KI-Einsatz",
                  len([c for c in geo_companies if c.get("kategorie") == "ECHTER_EINSATZ"]))

# ──────────────────────────────────────────────────────────
# PAGE: Neu analysieren
# ──────────────────────────────────────────────────────────
elif page == "➕ Neu analysieren":
    st.header("➕ Unternehmen hinzufügen & analysieren")

    with st.form("new_company_form"):
        st.subheader("Unternehmensdaten")
        col1, col2 = st.columns(2)

        with col1:
            name = st.text_input("Firmenname *", placeholder="Musterfirma GmbH")
            website = st.text_input("Website *", placeholder="https://musterfirma.de")
            industry = st.text_input("Branche", placeholder="IT / Logistik / Handel...")
        with col2:
            address = st.text_input("Straße & Hausnummer", placeholder="Hauptstraße 1")
            postal_code = st.text_input("PLZ", placeholder="23843")
            city = st.text_input("Stadt", placeholder="Bad Oldesloe")
            employee_count = st.selectbox("Mitarbeiter",
                ["–", "1-9", "10-49", "50-249", "250-999", "1000+"])

        col_opt1, col_opt2 = st.columns(2)
        with col_opt1:
            do_geo = st.checkbox("Geocodierung (Adresse → Koordinaten)", value=True)
        with col_opt2:
            do_bio = st.checkbox("Biografie generieren", value=True)

        submitted = st.form_submit_button("🚀 Analysieren", type="primary")

    if submitted:
        if not name or not website:
            st.error("Bitte Firmenname und Website angeben.")
        elif not os.getenv("OPENAI_API_KEY"):
            st.error("OpenAI API Key fehlt. Bitte in der Sidebar eingeben.")
        else:
            progress = st.progress(0, "Starte Analyse...")

            # 1. Scraping
            progress.progress(20, "🌐 Website wird gelesen...")
            scrape_result = scraper.scrape_website(website)

            if scrape_result["error"]:
                st.warning(f"Scraping-Warnung: {scrape_result['error']}")
                website_text = f"Firmenname: {name}"
            else:
                website_text = scrape_result["text"]
                st.success(f"✅ {scrape_result['pages_scraped']} Seiten gelesen · "
                           f"{len(scrape_result['keyword_hits'])} Keyword-Treffer: "
                           f"{', '.join(scrape_result['keyword_hits'][:5])}")

            # 2. Geocodierung
            lat, lng = None, None
            if do_geo and (address or city):
                progress.progress(40, "📍 Geocodierung...")
                lat, lng = geo_mapper.geocode_address(address, city, postal_code)
                if lat:
                    st.success(f"📍 Koordinaten: {lat:.4f}, {lng:.4f}")
                else:
                    st.info("Adresse konnte nicht geocodiert werden.")

            # 3. Company speichern
            company_id = db.upsert_company(
                name=name, website=website, address=address,
                city=city, postal_code=postal_code,
                lat=lat, lng=lng, industry=industry,
                employee_count=employee_count
            )

            # 4. LLM-Analyse
            progress.progress(60, "🤖 KI-Analyse...")
            try:
                classification = analyzer.classify_company(name, website_text)

                biografie = ""
                if do_bio:
                    progress.progress(80, "✍️ Biografie wird geschrieben...")
                    biografie = analyzer.generate_biography(name, website_text, classification)

                db.save_analysis(
                    company_id=company_id,
                    kategorie=classification["kategorie"],
                    begruendung=classification["begruendung"],
                    ki_anwendungen=classification["ki_anwendungen"],
                    vertrauen=classification["vertrauen"],
                    biografie=biografie,
                    raw_text=website_text[:2000]
                )

                db.log_event(company_id, "NEU",
                             f"Analysiert: {classification['kategorie']} (Score: {classification['vertrauen']})")

                progress.progress(100, "✅ Fertig!")

                # Ergebnis anzeigen
                st.markdown("---")
                st.subheader("📊 Analyseergebnis")

                col1, col2 = st.columns([2, 1])
                with col1:
                    st.markdown(f"### {name}")
                    kat = classification["kategorie"]
                    st.markdown(category_badge(kat), unsafe_allow_html=True)
                    st.markdown(f"\n**Begründung:** {classification['begruendung']}")

                    if classification.get("ki_anwendungen"):
                        st.markdown("**KI-Anwendungen:**")
                        for app in classification["ki_anwendungen"]:
                            st.markdown(f"• {app}")

                    if biografie:
                        st.markdown("---")
                        st.markdown("**Biografie:**")
                        st.markdown(f"*{biografie}*")

                with col2:
                    st.metric("Vertrauens-Score", f"{classification['vertrauen']}/100")

            except Exception as e:
                st.error(f"LLM-Analyse fehlgeschlagen: {e}")
                progress.empty()

# ──────────────────────────────────────────────────────────
# PAGE: Excel-Import
# ──────────────────────────────────────────────────────────
elif page == "📤 Excel-Import":
    st.header("📤 Massenimport aus Excel")

    # Vorlage herunterladen
    st.subheader("1. Excel-Vorlage herunterladen")
    st.markdown("Trage alle Stormarn-Unternehmen in die Vorlage ein und lade sie dann hoch.")

    # Vorlage erstellen
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    def create_template():
        wb = Workbook()
        ws = wb.active
        ws.title = "Unternehmen"
        headers = ["Firmenname*", "Website*", "Straße", "PLZ", "Stadt", "Branche", "Mitarbeiter", "Notizen"]
        header_fill = PatternFill("solid", start_color="1A5276")
        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        examples = [
            ["Musterfirma GmbH", "https://musterfirma.de", "Hauptstraße 1", "22941", "Bargteheide", "IT", "50-249", ""],
            ["Technik AG", "https://technik-ag.de", "Industrieweg 5", "21465", "Reinbek", "Maschinenbau", "250-999", ""],
        ]
        for row_idx, row_data in enumerate(examples, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        widths = [30, 35, 25, 8, 20, 20, 15, 25]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    template_bytes = create_template()
    st.download_button(
        label="⬇️ Vorlage herunterladen",
        data=template_bytes,
        file_name="Stormarn_Radar_Vorlage.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    st.markdown("---")

    # Upload
    st.subheader("2. Ausgefüllte Excel-Datei hochladen")
    uploaded_file = st.file_uploader(
        "Excel-Datei hier hochladen",
        type=["xlsx"],
        help="Nur .xlsx Dateien mit dem Vorlage-Format"
    )

    if uploaded_file:
        df, error = bulk_analyzer.read_excel(uploaded_file.read())

        if error:
            st.error(f"Fehler beim Lesen: {error}")
        else:
            st.success(f"✅ {len(df)} Unternehmen gefunden!")
            st.dataframe(df, use_container_width=True, hide_index=True)

            st.markdown("---")
            st.subheader("3. Analyse starten")

            col1, col2 = st.columns(2)
            with col1:
                do_geo = st.checkbox("Geocodierung", value=True)
            with col2:
                do_bio = st.checkbox("Biografien generieren", value=True)

            st.warning(f"⚠️ {len(df)} Unternehmen werden analysiert. Kosten: ca. {len(df) * 0.01:.2f}€")

            if not os.getenv("OPENAI_API_KEY"):
                st.error("OpenAI API Key fehlt. Bitte links unten eingeben.")
            else:
                if st.button("🚀 Alle analysieren", type="primary"):
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    results_placeholder = st.empty()

                    results = []

                    def update_progress(current, total, message):
                        progress_bar.progress(current / total)
                        status_text.text(message)

                    with st.spinner("Analyse läuft..."):
                        results = bulk_analyzer.analyze_batch(
                            df,
                            progress_callback=update_progress,
                            do_geo=do_geo,
                            do_bio=do_bio
                        )

                    progress_bar.progress(1.0)
                    status_text.text("✅ Fertig!")

                    # Ergebnisse anzeigen
                    success = [r for r in results if r["status"] == "success"]
                    errors = [r for r in results if r["status"] == "error"]

                    st.markdown("---")
                    st.subheader("📊 Ergebnisse")

                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("✅ Erfolgreich", len(success))
                    with col2:
                        st.metric("❌ Fehler", len(errors))
                    with col3:
                        echter = len([r for r in success if r.get("kategorie") == "ECHTER_EINSATZ"])
                        st.metric("🎯 Echter KI-Einsatz", echter)

                    if success:
                        results_df = pd.DataFrame(success)
                        results_df = results_df[["name", "website", "kategorie", "vertrauen"]]
                        results_df.columns = ["Unternehmen", "Website", "Kategorie", "Score"]
                        st.dataframe(results_df, use_container_width=True, hide_index=True)

                    if errors:
                        st.markdown("**Fehler:**")
                        for e in errors:
                            st.error(f"{e['name']}: {e['error']}")


# ──────────────────────────────────────────────────────────
# PAGE: Wirtschaftsdaten
# ──────────────────────────────────────────────────────────
elif page == "📊 Wirtschaftsdaten":
    st.header("📊 Wirtschaftsdaten Stormarn – 8.838 Unternehmen")
    st.markdown("Lade die offizielle Wirtschaftsdaten-Excel hoch und importiere alle Stormarn-Firmen direkt ins Radar.")

    uploaded = st.file_uploader(
        "Wirtschaftsdaten_Stormarn.xlsx hochladen",
        type=["xlsx"],
        help="Die Excel-Datei mit allen Stormarn-Unternehmen"
    )

    if uploaded:
        with st.spinner("Datei wird gelesen..."):
            df, error = wirtschaftsdaten_importer.load_wirtschaftsdaten(uploaded.read())

        if error:
            st.error(f"Fehler: {error}")
        else:
            stats = wirtschaftsdaten_importer.get_stats(df)

            # Statistiken
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("🏢 Unternehmen gesamt", f"{stats['total']:,}")
            with col2:
                st.metric("🌐 Mit Website", f"{stats['with_website']:,}")
            with col3:
                st.metric("📍 Ohne Website", f"{stats['without_website']:,}")

            st.markdown("---")

            # Filter
            st.subheader("🔧 Filter")
            col4, col5 = st.columns(2)
            with col4:
                only_web = st.checkbox("Nur Firmen mit Website", value=True)
                selected_cities = st.multiselect(
                    "Städte filtern",
                    list(stats["cities"].keys()),
                    default=[]
                )
            with col5:
                st.markdown("**Top Branchen:**")
                for b, count in list(stats["top_branches"].items())[:5]:
                    st.text(f"• {b[:50]}: {count}")

            # Gefilterte Daten
            filtered = wirtschaftsdaten_importer.filter_companies(
                df,
                only_with_website=only_web,
                cities=selected_cities if selected_cities else None
            )

            st.info(f"📋 **{len(filtered):,} Unternehmen** nach Filter")
            st.dataframe(
                filtered[["name", "ort", "website", "branche", "mitarbeiter"]].head(100),
                use_container_width=True,
                hide_index=True
            )

            if len(filtered) > 100:
                st.caption(f"Zeige erste 100 von {len(filtered):,} Unternehmen")

            st.markdown("---")
            st.subheader("🚀 Import & Analyse")

            col6, col7 = st.columns(2)
            with col6:
                do_import = st.button("💾 Alle in Datenbank speichern", type="secondary")
            with col7:
                do_analyze = st.button("🤖 Direkt KI-Analyse starten", type="primary",
                                       help="Analysiert alle Firmen MIT Website")

            if do_import:
                progress = st.progress(0)
                saved = 0
                for i, row in filtered.iterrows():
                    try:
                        db.upsert_company(
                            name=row["name"],
                            website=row.get("website", ""),
                            address=row.get("adresse", ""),
                            city=row.get("ort", ""),
                            postal_code=row.get("plz", ""),
                            industry=row.get("branche", ""),
                            employee_count=str(row.get("mitarbeiter", ""))
                        )
                        saved += 1
                        if saved % 100 == 0:
                            progress.progress(saved / len(filtered))
                    except Exception:
                        continue
                progress.progress(1.0)
                st.success(f"✅ {saved:,} Unternehmen gespeichert!")
                db.log_event(None, "WIRTSCHAFTSDATEN_IMPORT",
                            f"{saved} Firmen aus Wirtschaftsdaten importiert")

            if do_analyze:
                if not os.getenv("OPENAI_API_KEY"):
                    st.error("OpenAI API Key fehlt!")
                else:
                    with_web = filtered[filtered["website"] != ""].copy()
                    st.info(f"Analysiere {len(with_web):,} Firmen mit Website...")

                    # In bulk_analyzer Format bringen
                    analyze_df = pd.DataFrame({
                        "Firmenname*": with_web["name"].values,
                        "Website*": with_web["website"].values,
                        "Straße": with_web["adresse"].values,
                        "PLZ": with_web["plz"].values,
                        "Stadt": with_web["ort"].values,
                        "Branche": with_web["branche"].values,
                        "Mitarbeiter": with_web["mitarbeiter"].values,
                        "Notizen": ""
                    })

                    ana_progress = st.progress(0)
                    ana_status = st.empty()

                    def ana_cb(current, total, message):
                        ana_progress.progress(current / total)
                        ana_status.text(message)

                    results = bulk_analyzer.analyze_batch(
                        analyze_df,
                        progress_callback=ana_cb,
                        do_geo=True
                    )
                    success = [r for r in results if r["status"] == "success"]
                    st.success(f"✅ {len(success):,} Firmen analysiert!")


# ──────────────────────────────────────────────────────────
# PAGE: Auto-Suche
# ──────────────────────────────────────────────────────────
elif page == "🔍 Auto-Suche":
    st.header("🔍 Automatische Firmen-Suche in Stormarn")
    st.markdown("Sucht automatisch Firmen aus Gelbe Seiten, Handelsregister und Wer-zu-Wem.")

    col1, col2 = st.columns(2)

    with col1:
        st.subheader("Quellen")
        use_gelbe = st.checkbox("Gelbe Seiten", value=True)
        use_wer_zu_wem = st.checkbox("Wer-zu-Wem", value=True)
        use_handelsregister = st.checkbox("Handelsregister", value=False,
                                          help="Langsamer aber offiziell")

    with col2:
        st.subheader("Städte")
        all_cities = company_finder.STORMARN_CITIES
        selected_cities = st.multiselect(
            "Städte auswählen",
            all_cities,
            default=all_cities[:5]
        )

    st.markdown("---")
    col3, col4 = st.columns(2)
    with col3:
        auto_analyze = st.checkbox("Direkt KI-Analyse starten", value=False)
    with col4:
        do_geo = st.checkbox("Geocodierung", value=True)

    if selected_cities:
        est_time = len(selected_cities) * 3
        st.info(f"⏱️ Geschätzte Zeit: ca. {est_time} Minuten")

        if st.button("🚀 Suche starten", type="primary"):
            sources = []
            if use_gelbe: sources.append("gelbe_seiten")
            if use_wer_zu_wem: sources.append("wer_zu_wem")
            if use_handelsregister: sources.append("handelsregister")

            if not sources:
                st.error("Bitte mindestens eine Quelle auswählen.")
                st.stop()

            status = st.empty()
            progress_bar = st.progress(0)
            total_steps = len(selected_cities) * len(sources)
            step = [0]

            def update_status(message):
                step[0] += 1
                progress_bar.progress(min(step[0] / total_steps, 1.0))
                status.text(message)

            with st.spinner("Suche läuft..."):
                found = company_finder.find_companies_in_stormarn(
                    cities=selected_cities,
                    sources=sources,
                    progress_callback=update_status
                )

            progress_bar.progress(1.0)
            status.text(f"✅ {len(found)} Firmen gefunden!")

            if found:
                st.markdown("---")
                st.subheader(f"📋 {len(found)} gefundene Firmen")
                found_df = pd.DataFrame(found)
                display_cols = [c for c in ["name", "city", "website", "source"] if c in found_df.columns]
                st.dataframe(found_df[display_cols], use_container_width=True, hide_index=True)

                if st.button("💾 Alle in Datenbank speichern"):
                    saved = 0
                    for c in found:
                        try:
                            db.upsert_company(
                                name=c["name"], website=c.get("website", ""),
                                address=c.get("address", ""), city=c.get("city", ""),
                                postal_code=c.get("postal_code", "")
                            )
                            saved += 1
                        except Exception:
                            continue
                    st.success(f"✅ {saved} Firmen gespeichert!")
            else:
                st.warning("Keine Firmen gefunden. Versuche andere Städte oder Quellen.")



# ──────────────────────────────────────────────────────────
# PAGE: KI-Ranking
# ──────────────────────────────────────────────────────────
elif page == "🏅 KI-Ranking":
    st.header("🏅 KI-Reifegrad Ranking")
    st.markdown("Alle analysierten Stormarn-Unternehmen nach KI-Score (1–10) bewertet.")

    companies = db.get_all_companies()
    analyzed = [c for c in companies if c.get("kategorie")]

    if not analyzed:
        st.info("Noch keine Unternehmen analysiert.")
    else:
        # Scores berechnen
        scored = []
        for c in analyzed:
            score_data = ki_scorer.calculate_ki_score(
                kategorie=c.get("kategorie", "KEIN_KI"),
                vertrauen=c.get("vertrauen", 50),
                ki_anwendungen=c.get("ki_anwendungen", []),
                raw_text=c.get("raw_text", "")
            )
            scored.append({**c, **score_data})

        scored = sorted(scored, key=lambda x: x["score"], reverse=True)

        # Top 3 Podium
        st.subheader("🏆 Top 3 KI-Vorreiter")
        top3 = scored[:3]
        cols = st.columns(3)
        medals = ["🥇", "🥈", "🥉"]
        for i, (col, firm) in enumerate(zip(cols, top3)):
            with col:
                st.metric(
                    f"{medals[i]} {firm['name'][:20]}",
                    f"{firm['badge']} {firm['score']}/10",
                    firm["level"]
                )

        st.markdown("---")

        # Vollständiges Ranking
        st.subheader("📋 Vollständiges Ranking")

        score_filter = st.slider("Mindest-Score", 1, 10, 1)
        filtered = [c for c in scored if c["score"] >= score_filter]

        for rank, firm in enumerate(filtered, 1):
            with st.expander(f"#{rank} {firm['badge']} {firm['name']} – Score: {firm['score']}/10 · {firm['level']}"):
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("KI-Score", f"{firm['score']}/10")
                with col2:
                    st.metric("Kategorie", firm.get("kategorie", "").replace("_", " "))
                with col3:
                    st.metric("Vertrauen", f"{firm.get('vertrauen', 0)}%")
                st.caption(firm.get("erklaerung", ""))
                if firm.get("website"):
                    st.markdown(f"🌐 [{firm['website']}]({firm['website']})")


# ──────────────────────────────────────────────────────────
# PAGE: News-Monitor
# ──────────────────────────────────────────────────────────
elif page == "📰 News-Monitor":
    st.header("📰 News-Monitor")
    st.markdown("Aktuelle KI-News aus Stormarn und von deinen analysierten Unternehmen.")

    tab1, tab2 = st.tabs(["🌍 Stormarn KI-News", "🏢 Unternehmens-News"])

    with tab1:
        if st.button("🔄 News laden"):
            with st.spinner("Suche aktuelle News..."):
                news = news_monitor.search_stormarn_ki_news()

            if news:
                for item in news:
                    with st.container():
                        st.markdown(f"**[{item['title']}]({item['link']})**")
                        st.caption(f"📅 {item.get('published', '')} · {item.get('source', '')}")
                        st.markdown("---")
            else:
                st.info("Keine aktuellen News gefunden.")

    with tab2:
        companies = db.get_all_companies()
        analyzed = [c for c in companies if c.get("kategorie") == "ECHTER_EINSATZ"]

        if not analyzed:
            st.info("Erst Unternehmen analysieren um deren News zu sehen.")
        else:
            selected = st.selectbox(
                "Unternehmen auswählen",
                [c["name"] for c in analyzed[:20]]
            )

            if st.button("🔍 News suchen"):
                with st.spinner(f"Suche News für {selected}..."):
                    news = news_monitor.search_company_news(selected)

                if news:
                    for item in news:
                        st.markdown(f"**[{item['title']}]({item['link']})**")
                        ki_badge = "🤖 KI-relevant" if item.get("ki_relevant") else ""
                        st.caption(f"📅 {item.get('published', '')} · {item.get('source', '')} {ki_badge}")
                        st.markdown("---")
                else:
                    st.info("Keine News gefunden.")


# ──────────────────────────────────────────────────────────
# PAGE: Regionalvergleich
# ──────────────────────────────────────────────────────────
elif page == "🌍 Regionalvergleich":
    st.header("🌍 Regionalvergleich – Stormarn vs. Schleswig-Holstein")

    companies = db.get_all_companies()
    analyzed = [c for c in companies if c.get("kategorie")]
    ki_quote = regional_compare.get_stormarn_ki_quote(analyzed)

    compare_data = regional_compare.get_comparison_data(
        actual_ki_quote=ki_quote if analyzed else None
    )

    metric = st.selectbox(
        "Vergleichskriterium",
        ["ki_quote_est", "digitalquote", "breitband", "startup_index"],
        format_func=lambda x: {
            "ki_quote_est": "KI-Quote (%)",
            "digitalquote": "Digitalisierungsquote (%)",
            "breitband": "Breitbandversorgung (%)",
            "startup_index": "Startup-Index (0-100)"
        }[x]
    )

    ranking = regional_compare.get_ranking(metric, compare_data)
    position = regional_compare.get_stormarn_position(metric, compare_data)

    # Position anzeigen
    st.metric(
        f"Stormarn Platz",
        f"{position['position']} von {position['total']}",
        f"Besser als {position['besser_als']} Kreise"
    )

    if analyzed:
        st.success(f"🎯 Echte KI-Quote aus {len(analyzed)} analysierten Firmen: **{ki_quote}%**")

    st.markdown("---")
    st.subheader("📊 Ranking")

    for i, (kreis, wert) in enumerate(ranking, 1):
        farbe = compare_data[kreis].get("farbe", "#3498DB")
        is_stormarn = kreis == "Kreis Stormarn"
        prefix = "👉 " if is_stormarn else ""
        bold_start = "**" if is_stormarn else ""
        bold_end = "**" if is_stormarn else ""
        st.markdown(f"{prefix}{i}. {bold_start}{kreis}{bold_end}: **{wert}{'%' if 'quote' in metric or metric == 'breitband' else ''}**")


# ──────────────────────────────────────────────────────────
# PAGE: IHK-Bericht
# ──────────────────────────────────────────────────────────
elif page == "📑 IHK-Bericht":
    st.header("📑 IHK-Bericht erstellen")
    st.markdown("Erstelle einen professionellen PDF-Bericht für die IHK oder Kreisverwaltung.")

    companies = db.get_all_companies()
    analyzed = [c for c in companies if c.get("kategorie")]

    if not analyzed:
        st.warning("Noch keine Unternehmen analysiert. Bitte erst Firmen analysieren.")
    else:
        echter = len([c for c in analyzed if c.get("kategorie") == "ECHTER_EINSATZ"])
        integration = len([c for c in analyzed if c.get("kategorie") == "INTEGRATION"])
        kein_ki = len([c for c in analyzed if c.get("kategorie") == "KEIN_KI"])

        col1, col2, col3, col4 = st.columns(4)
        with col1: st.metric("Analysiert", len(analyzed))
        with col2: st.metric("✅ Echter Einsatz", echter)
        with col3: st.metric("🔗 Integration", integration)
        with col4: st.metric("❌ Kein KI", kein_ki)

        st.markdown("---")

        report_title = st.text_input(
            "Berichtstitel",
            value=f"KI-Radar Kreis Stormarn – {datetime.now().strftime('%B %Y')}"
        )

        if st.button("📄 PDF-Bericht erstellen", type="primary"):
            with st.spinner("Bericht wird erstellt..."):
                stats = {
                    "total": len(companies),
                    "analyzed": len(analyzed),
                    "echter_einsatz": echter,
                    "integration": integration,
                    "kein_ki": kein_ki
                }
                pdf_bytes = report_generator.generate_ihk_report(
                    companies=analyzed,
                    stats=stats,
                    title=report_title
                )

            st.download_button(
                label="⬇️ IHK-Bericht herunterladen",
                data=pdf_bytes,
                file_name=f"Stormarn_KI_Radar_{datetime.now().strftime('%Y%m%d')}.pdf",
                mime="application/pdf"
            )
            st.success("✅ Bericht fertig!")


# ──────────────────────────────────────────────────────────
# PAGE: Job-Radar
# ──────────────────────────────────────────────────────────
elif page == "💼 Job-Radar":
    st.header("💼 Job-Radar – KI-Stellenanzeigen als Indikator")
    st.markdown("KI-Stellenanzeigen sind der stärkste Beweis für echten KI-Einsatz in einem Unternehmen.")

    tab1, tab2 = st.tabs(["🔍 Einzelne Firma analysieren", "📊 Alle KI-Firmen prüfen"])

    with tab1:
        companies = db.get_all_companies()
        company_names = [c["name"] for c in companies if c.get("name")]

        if company_names:
            selected = st.selectbox("Unternehmen auswählen", company_names)
            sel_company = next((c for c in companies if c["name"] == selected), {})

            col1, col2 = st.columns(2)
            with col1:
                st.text_input("Website", value=sel_company.get("website", ""), key="job_website")
            with col2:
                st.text_input("Firmenname", value=selected, key="job_name")

            if st.button("🔍 KI-Jobs suchen", type="primary"):
                website = st.session_state.get("job_website", sel_company.get("website", ""))
                with st.spinner(f"Suche KI-Stellenanzeigen für {selected}..."):
                    result = job_radar.analyze_company_jobs(selected, website)

                st.markdown("---")
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("KI-Jobs gefunden", result["ki_job_count"])
                with col2:
                    st.metric("Max. KI-Score", f"{result['ki_job_score']}/10")
                with col3:
                    st.metric("Signal", "✅ Ja" if result["has_ki_jobs"] else "❌ Nein")

                st.info(result["signal_strength"])

                if result["jobs"]:
                    st.subheader("Gefundene KI-Stellen:")
                    for job in result["jobs"]:
                        with st.expander(f"💼 {job['title']} (Score: {job['ki_score']}/10)"):
                            st.markdown(f"**Quelle:** {job['source']}")
                            if job.get("link"):
                                st.markdown(f"**Link:** [{job['link']}]({job['link']})")
                            st.markdown(f"**KI-Signal:** `{job.get('ki_signal', '')}`")
                else:
                    st.info("Keine KI-Stellenanzeigen gefunden. Das bedeutet nicht zwingend kein KI-Einsatz!")
        else:
            st.info("Erst Unternehmen in die Datenbank importieren.")

    with tab2:
        st.markdown("Prüft alle analysierten KI-Firmen auf Stellenanzeigen – als Qualitätsprüfung.")
        ki_companies = [c for c in db.get_all_companies()
                        if c.get("kategorie") in ("ECHTER_EINSATZ", "INTEGRATION")]

        st.info(f"**{len(ki_companies)}** KI-Firmen können geprüft werden.")
        st.warning("⚠️ Das dauert ca. 2 Minuten pro Firma. Wähle eine kleine Gruppe!")

        max_check = st.slider("Maximale Anzahl prüfen", 1, min(20, len(ki_companies)), 5)

        if st.button("🚀 Job-Check starten"):
            progress = st.progress(0)
            results = []

            for i, comp in enumerate(ki_companies[:max_check]):
                progress.progress((i + 1) / max_check)
                result = job_radar.analyze_company_jobs(
                    comp["name"], comp.get("website", "")
                )
                results.append({**comp, "job_result": result})

            progress.progress(1.0)

            with_jobs = [r for r in results if r["job_result"]["has_ki_jobs"]]
            st.success(f"✅ {len(with_jobs)} von {len(results)} Firmen haben KI-Stellenanzeigen!")

            for r in results:
                jr = r["job_result"]
                icon = "✅" if jr["has_ki_jobs"] else "❌"
                st.markdown(
                    f"{icon} **{r['name']}** – {jr['ki_job_count']} KI-Jobs · "
                    f"Score: {jr['ki_job_score']}/10 · {jr['signal_strength']}"
                )



# ──────────────────────────────────────────────────────────
# PAGE: Re-Analyse
# ──────────────────────────────────────────────────────────
elif page == "🔄 Re-Analyse":
    st.header("🔄 Qualitäts-Verbesserung & Auto-Refresh")
    st.markdown("Verbessert die Analyse-Qualität durch tieferes Scanning und aktualisiert veraltete Ergebnisse.")

    # Statistiken
    refresh_stats = reanalyzer.get_refresh_stats()
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Analysiert gesamt", refresh_stats["total_analyzed"])
    with col2:
        st.metric("⚠️ Unsichere Ergebnisse", refresh_stats["uncertain_count"],
                  help=f"Vertrauen unter {reanalyzer.LOW_CONFIDENCE_THRESHOLD}%")
    with col3:
        st.metric("⏰ Veraltete Analysen", refresh_stats["stale_count"],
                  help=f"Älter als {reanalyzer.REFRESH_DAYS} Tage")
    with col4:
        st.metric("🔄 Brauchen Review", refresh_stats["needs_review"])

    st.markdown("---")

    tab1, tab2, tab3 = st.tabs(["⚠️ Unsichere re-analysieren",
                                 "⏰ Veraltete aktualisieren",
                                 "🎯 Einzelne Firma"])

    with tab1:
        st.subheader("Unsichere Ergebnisse verbessern")
        st.markdown(f"Firmen mit Vertrauen unter **{reanalyzer.LOW_CONFIDENCE_THRESHOLD}%** werden tiefer gescannt.")

        threshold = st.slider("Vertrauen-Schwellwert (%)", 30, 80,
                               reanalyzer.LOW_CONFIDENCE_THRESHOLD)
        uncertain = reanalyzer.get_uncertain_companies(threshold)

        if uncertain:
            st.info(f"**{len(uncertain)}** Firmen werden neu analysiert – "
                    f"ca. {len(uncertain) * 20} Sekunden")

            if st.button("🚀 Re-Analyse starten", type="primary", key="reanalyze_uncertain"):
                progress = st.progress(0)
                status = st.empty()
                results_placeholder = st.empty()

                def cb(current, total, msg):
                    progress.progress(current / total)
                    status.text(msg)

                results = reanalyzer.run_reanalysis_batch(uncertain, cb)

                progress.progress(1.0)
                changed = [r for r in results if r.get("changed")]
                improved = [r for r in results if r.get("improved")]
                errors = [r for r in results if r.get("error")]

                st.success(f"✅ {len(results)} Firmen re-analysiert!")
                col1, col2, col3 = st.columns(3)
                with col1: st.metric("Verändert", len(changed))
                with col2: st.metric("Verbessert", len(improved))
                with col3: st.metric("Fehler", len(errors))

                if changed:
                    st.subheader("🔄 Veränderte Klassifizierungen:")
                    for r in changed:
                        old_k = r.get("old_kategorie", "?")
                        new_k = r.get("new_kategorie", "?")
                        st.markdown(
                            f"**{r['company']}**: "
                            f"`{old_k}` ({r.get('old_vertrauen',0)}%) → "
                            f"`{new_k}` ({r.get('new_vertrauen',0)}%)"
                        )
        else:
            st.success("✅ Alle Analysen haben ausreichendes Vertrauen!")

    with tab2:
        st.subheader("Veraltete Analysen aktualisieren")
        days = st.slider("Analysen älter als X Tage", 7, 90, reanalyzer.REFRESH_DAYS)
        stale = reanalyzer.get_stale_companies(days)

        if stale:
            st.info(f"**{len(stale)}** Firmen seit über {days} Tagen nicht analysiert")
            max_refresh = st.slider("Maximal aktualisieren", 1,
                                    min(50, len(stale)), min(10, len(stale)))

            if st.button("🔄 Aktualisierung starten", type="primary", key="refresh_stale"):
                progress = st.progress(0)
                status = st.empty()

                def cb2(current, total, msg):
                    progress.progress(current / total)
                    status.text(msg)

                results = reanalyzer.run_reanalysis_batch(stale[:max_refresh], cb2)
                progress.progress(1.0)
                changed = [r for r in results if r.get("changed")]
                st.success(f"✅ {len(results)} Firmen aktualisiert, {len(changed)} verändert!")
        else:
            st.success(f"✅ Alle Analysen sind aktuell (jünger als {days} Tage)!")

    with tab3:
        st.subheader("Einzelne Firma neu analysieren")
        companies = db.get_all_companies()
        company_names = [c["name"] for c in companies if c.get("website")]

        if company_names:
            selected_name = st.selectbox("Firma auswählen", company_names)
            sel = next((c for c in companies if c["name"] == selected_name), None)

            if sel:
                col1, col2 = st.columns(2)
                with col1:
                    st.metric("Aktuell", sel.get("kategorie", "Nicht analysiert"))
                with col2:
                    st.metric("Vertrauen", f"{sel.get('vertrauen', 0)}%")

                deep_scan = st.checkbox("Tiefes Scanning (mehr Unterseiten)", value=True)

                if st.button("🔍 Jetzt re-analysieren", type="primary"):
                    with st.spinner(f"Analysiere {selected_name} tief..."):
                        result = reanalyzer.reanalyze_company(sel, deep=deep_scan)

                    if result.get("error"):
                        st.error(f"Fehler: {result['error']}")
                    else:
                        if result["changed"]:
                            st.warning(
                                f"⚠️ Klassifizierung geändert: "
                                f"`{result['old_kategorie']}` → `{result['new_kategorie']}`"
                            )
                        else:
                            st.success(
                                f"✅ Bestätigt: `{result['new_kategorie']}` "
                                f"({result['new_vertrauen']}% Vertrauen)"
                            )

                        col1, col2 = st.columns(2)
                        with col1:
                            st.metric("Altes Vertrauen", f"{result['old_vertrauen']}%")
                        with col2:
                            delta = result['new_vertrauen'] - result['old_vertrauen']
                            st.metric("Neues Vertrauen", f"{result['new_vertrauen']}%",
                                     delta=f"{delta:+d}%")



# ──────────────────────────────────────────────────────────
# PAGE: Wochenbericht
# ──────────────────────────────────────────────────────────
elif page == "📧 Wochenbericht":
    st.header("📧 Automatischer Wochenbericht")
    st.markdown("Jeden Montag automatisch einen Bericht per E-Mail erhalten.")

    tab1, tab2 = st.tabs(["📬 Bericht senden", "👁️ Vorschau"])

    with tab1:
        st.subheader("E-Mail Konfiguration")

        smtp_configured = bool(os.getenv("SMTP_USER") or os.getenv("SMTP_PASSWORD"))

        if not smtp_configured:
            st.warning("⚠️ SMTP noch nicht konfiguriert. Gehe zu **Einstellungen** → E-Mail Alerts.")
            st.info("Füge in Streamlit Secrets ein:\n```\nSMTP_USER = 'deine@gmail.com'\nSMTP_PASSWORD = 'app-passwort'\nREPORT_EMAIL = 'empfaenger@example.com'\n```")
        else:
            st.success("✅ SMTP konfiguriert")

        recipient = st.text_input(
            "Empfänger-E-Mail",
            value=os.getenv("REPORT_EMAIL", ""),
            placeholder="empfaenger@example.com"
        )

        col1, col2 = st.columns(2)
        with col1:
            if st.button("📤 Bericht jetzt senden", type="primary"):
                if not recipient:
                    st.error("Bitte Empfänger-E-Mail eingeben!")
                else:
                    with st.spinner("Bericht wird gesendet..."):
                        result = weekly_report.send_weekly_report(recipient)

                    if result["success"]:
                        st.success(f"✅ {result['message']}")
                    else:
                        st.error(f"❌ {result['message']}")
        with col2:
            st.info("💡 Für automatischen Montags-Versand: Nutze einen Cron-Job oder Scheduler")

    with tab2:
        st.subheader("Berichts-Vorschau")
        if st.button("🔄 Vorschau generieren"):
            with st.spinner("Generiere Vorschau..."):
                html_preview = weekly_report.get_report_preview()
            st.components.v1.html(html_preview, height=800, scrolling=True)


# ──────────────────────────────────────────────────────────
# ──────────────────────────────────────────────────────────
    st.header("📈 Trend-Analyse")

    companies = db.get_all_companies()
    analyzed = [c for c in companies if c.get("kategorie")]

    if not analyzed:
        st.info("Noch zu wenig Daten für eine Trend-Analyse.")
        st.stop()

    # Verteilungs-Chart
    st.subheader("Kategorie-Verteilung")
    stats = db.get_stats()
    by_cat = stats["by_category"]

    chart_data = pd.DataFrame({
        "Kategorie": [CATEGORY_LABELS.get(k, k) for k in by_cat.keys()],
        "Anzahl": list(by_cat.values())
    })
    st.bar_chart(chart_data.set_index("Kategorie"))

    # Branchen-Übersicht
    st.subheader("Branchen-Verteilung")
    industries = {}
    for c in analyzed:
        ind = c.get("industry", "Unbekannt") or "Unbekannt"
        industries[ind] = industries.get(ind, 0) + 1

    if industries:
        ind_df = pd.DataFrame(
            {"Branche": list(industries.keys()), "Anzahl": list(industries.values())}
        ).sort_values("Anzahl", ascending=False)
        st.bar_chart(ind_df.set_index("Branche"))

    # LLM-Trend-Report
    st.subheader("🤖 KI-Trend-Report")
    existing = db.get_latest_trend_report()
    if existing:
        st.info(f"Letzter Report: {existing['created_at'][:16]}")
        st.markdown(existing["report_text"])

    if st.button("Neuen Trend-Report generieren"):
        if not os.getenv("OPENAI_API_KEY"):
            st.error("OpenAI API Key fehlt.")
        else:
            with st.spinner("Analysiere Trends..."):
                report = analyzer.generate_trend_report(analyzed)
                db.save_trend_report(report, len(analyzed))
                st.markdown(report)
                st.success("Report gespeichert!")


# ──────────────────────────────────────────────────────────
# PAGE: Qualitäts-Check
# ──────────────────────────────────────────────────────────
elif page == "🔬 Qualitäts-Check":
    st.header("🔬 Qualitäts-Check & Re-Analyse")
    st.markdown("Verbessert die Analysequalität durch Tiefenscans und hält Daten aktuell.")

    companies = db.get_all_companies()
    stats = reanalyzer.get_freshness_stats(companies)

    # Übersicht
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("🏢 Gesamt", stats["total"])
    with col2:
        st.metric("✅ Analysiert", stats["analyzed"])
    with col3:
        delta_color = "normal" if stats["uncertain"] == 0 else "inverse"
        st.metric("⚠️ Unsicher", stats["uncertain"],
                  help="Vertrauen < 50% oder unbekannte Kategorie")
    with col4:
        st.metric("🕐 Veraltet (30d)", stats["stale_30"],
                  help="Seit mehr als 30 Tagen nicht analysiert")

    # Frische-Balken
    if stats["total"] > 0:
        st.markdown(f"**Daten-Aktualität:** {stats['fresh_percent']}% aktuell")
        st.progress(stats["fresh_percent"] / 100)

    st.markdown("---")

    tab1, tab2, tab3 = st.tabs(["🔬 Zweiter Durchlauf", "🔄 Veraltete aktualisieren", "📋 Details"])

    with tab1:
        st.subheader("🔬 Zweiter Analyse-Durchlauf")
        st.markdown(
            f"**{stats['uncertain']} Firmen** haben Vertrauen unter 50% oder unbekannte Kategorie. "
            "Der zweite Durchlauf scannt mehr Unterseiten für bessere Ergebnisse."
        )

        uncertain = reanalyzer.get_uncertain_companies(companies)
        if not uncertain:
            st.success("✅ Alle analysierten Firmen haben gutes Vertrauen!")
        else:
            st.warning(f"⚠️ {len(uncertain)} Firmen brauchen einen zweiten Durchlauf")

            # Vorschau der unsicheren Firmen
            with st.expander(f"Liste der {len(uncertain)} unsicheren Firmen"):
                for c in uncertain[:20]:
                    st.markdown(
                        f"• **{c['name']}** – Vertrauen: {c.get('vertrauen', 0)}% · "
                        f"Kategorie: {c.get('kategorie', 'unbekannt')}"
                    )

            est_cost = len(uncertain) * 0.01
            st.info(f"⏱️ Geschätzte Zeit: {len(uncertain) * 15} Sek. · Kosten: ca. €{est_cost:.2f}")

            if st.button("🔬 Zweiten Durchlauf starten", type="primary"):
                if not os.getenv("OPENAI_API_KEY"):
                    st.error("OpenAI API Key fehlt!")
                else:
                    progress = st.progress(0)
                    status = st.empty()

                    def cb(current, total, msg):
                        progress.progress(current / total)
                        status.text(msg)

                    with st.spinner("Tiefenanalyse läuft..."):
                        results = reanalyzer.run_second_pass(progress_callback=cb)

                    summary = reanalyzer.get_changes_summary(results)
                    progress.progress(1.0)
                    st.success(f"✅ {summary['successful']} Firmen neu analysiert!")

                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Verbessert", summary["improved"])
                    with col2:
                        st.metric("Kategorie geändert", summary["category_changed"])
                    with col3:
                        st.metric("Fehlgeschlagen", summary["failed"])

                    if summary["changes"]:
                        st.subheader("🔄 Kategorie-Änderungen:")
                        for ch in summary["changes"]:
                            st.markdown(
                                f"• **{ch['company']}**: "
                                f"`{ch['old_kategorie']}` → `{ch['new_kategorie']}` · "
                                f"Vertrauen: {ch['old_vertrauen']}% → {ch['new_vertrauen']}% · "
                                f"({ch.get('pages_scanned', 0)} Seiten gescannt)"
                            )

    with tab2:
        st.subheader("🔄 Veraltete Daten aktualisieren")
        days = st.slider("Firmen älter als X Tage neu analysieren", 7, 90, 30)
        stale = reanalyzer.get_stale_companies(companies, days=days)

        if not stale:
            st.success(f"✅ Alle Firmen wurden in den letzten {days} Tagen analysiert!")
        else:
            st.info(f"**{len(stale)}** Firmen wurden seit mehr als {days} Tagen nicht analysiert.")
            est = len(stale) * 0.01
            st.warning(f"⏱️ Geschätzte Zeit: {len(stale) * 15} Sek. · Kosten: ca. €{est:.2f}")

            if st.button(f"🔄 {len(stale)} Firmen aktualisieren", type="primary"):
                if not os.getenv("OPENAI_API_KEY"):
                    st.error("OpenAI API Key fehlt!")
                else:
                    progress = st.progress(0)
                    status = st.empty()

                    def refresh_cb(current, total, msg):
                        progress.progress(current / total)
                        status.text(msg)

                    with st.spinner("Aktualisierung läuft..."):
                        results = reanalyzer.refresh_stale_companies(
                            days=days, progress_callback=refresh_cb
                        )

                    summary = reanalyzer.get_changes_summary(results)
                    progress.progress(1.0)
                    st.success(f"✅ {summary['successful']} Firmen aktualisiert!")
                    st.metric("Kategorie-Änderungen", summary["category_changed"])

    with tab3:
        st.subheader("📋 Unterseiten-Scan Details")
        st.markdown("Welche Unterseiten wurden bei der letzten Analyse gescannt?")

        analyzed = [c for c in companies if c.get("kategorie")]
        if analyzed:
            selected = st.selectbox("Firma auswählen", [c["name"] for c in analyzed[:50]])
            sel = next((c for c in analyzed if c["name"] == selected), {})
            if sel:
                col1, col2 = st.columns(2)
                with col1:
                    st.metric("Vertrauen", f"{sel.get('vertrauen', 0)}%")
                with col2:
                    st.metric("Kategorie", sel.get("kategorie", "–"))

                subpages = sel.get("subpages", [])
                if subpages:
                    st.markdown("**Gescannte Unterseiten:**")
                    for p in subpages:
                        st.markdown(f"• [{p}]({p})")
                else:
                    st.info("Nur Hauptseite gescannt – Re-Analyse für mehr Details.")

                if st.button("🔬 Jetzt tief analysieren"):
                    with st.spinner(f"Analysiere {selected}..."):
                        result = reanalyzer.reanalyze_company(sel)
                    if result["success"]:
                        st.success(
                            f"✅ Fertig! Vertrauen: {result['old_vertrauen']}% → "
                            f"**{result['new_vertrauen']}%** · "
                            f"{result['pages_scanned']} Seiten gescannt"
                        )
                        if result["old_kategorie"] != result["new_kategorie"]:
                            st.info(
                                f"Kategorie geändert: `{result['old_kategorie']}` → "
                                f"`{result['new_kategorie']}`"
                            )
                    else:
                        st.error(f"Fehler: {result.get('error')}")


# ──────────────────────────────────────────────────────────
# PAGE: Aktivitätslog
# ──────────────────────────────────────────────────────────
elif page == "📋 Aktivitätslog":
    st.header("📋 Aktivitäts-Log")

    events = db.get_recent_events(100)

    if not events:
        st.info("Noch keine Aktivitäten")
    else:
        # DataFrame
        df = pd.DataFrame(events)
        df = df[["created_at", "company_name", "event_type", "message", "alerted"]]
        df.columns = ["Zeitpunkt", "Unternehmen", "Typ", "Details", "Gemeldet"]
        df["Gemeldet"] = df["Gemeldet"].map({0: "❌", 1: "✅"})
        st.dataframe(df, use_container_width=True, hide_index=True)

    # Alert-Bereich
    st.markdown("---")
    st.subheader("📧 E-Mail-Alert")
    alert_enabled = cfg.get("radar.alerts.enabled", False)

    if not alert_enabled:
        st.warning("Alerts sind in der config.yaml deaktiviert.")
    else:
        from_email = cfg.get("radar.alerts.from_email", "")
        to_email = cfg.get("radar.alerts.to_email", "")

        if not from_email or not to_email:
            st.warning("E-Mail nicht konfiguriert. Bitte `from_email` und `to_email` in config.yaml setzen.")
        else:
            st.info(f"Von: {from_email} → An: {to_email}")
            smtp_password = st.text_input("SMTP-Passwort (Gmail App-Passwort)", type="password")

            if st.button("🔔 Alert jetzt senden"):
                result = alert.check_and_alert(smtp_password)
                if result["status"] == "sent":
                    st.success(f"Alert gesendet: {result['count']} Ereignisse")
                elif result["status"] == "no_new_events":
                    st.info("Keine neuen Ereignisse")
                else:
                    st.warning(f"Status: {result['status']}")

# ──────────────────────────────────────────────────────────
# PAGE: PDF-Export
# ──────────────────────────────────────────────────────────
elif page == "📄 PDF-Export":
    st.header("📄 PDF-Steckbriefe exportieren")

    companies = db.get_all_companies()
    analyzed = [c for c in companies if c.get("kategorie")]

    if not analyzed:
        st.info("Noch keine analysierten Unternehmen.")
        st.stop()

    tab1, tab2 = st.tabs(["Einzelner Steckbrief", "Übersichts-PDF"])

    with tab1:
        company_names = {c["name"]: c["id"] for c in analyzed}
        selected_name = st.selectbox("Unternehmen wählen", list(company_names.keys()))

        if st.button("📄 Steckbrief generieren"):
            company_id = company_names[selected_name]
            company = db.get_company_by_id(company_id)
            analyses = db.get_analyses_for_company(company_id)

            if not analyses:
                st.error("Keine Analyse vorhanden.")
            else:
                analysis = analyses[0]
                with st.spinner("PDF wird erstellt..."):
                    path = pdf_export.generate_company_profile(company, analysis)
                with open(path, "rb") as f:
                    st.download_button(
                        label=f"⬇️ {selected_name} – Steckbrief.pdf",
                        data=f,
                        file_name=f"Steckbrief_{selected_name.replace(' ','_')}.pdf",
                        mime="application/pdf"
                    )

    with tab2:
        st.write(f"{len(analyzed)} Unternehmen werden ins Übersichts-PDF aufgenommen.")
        cat_filter = st.multiselect(
            "Nur diese Kategorien:",
            list(CATEGORY_LABELS.values()),
            default=list(CATEGORY_LABELS.values())
        )

        if st.button("📊 Übersichts-PDF generieren"):
            cat_keys = {v: k for k, v in CATEGORY_LABELS.items()}
            selected_keys = [cat_keys[c] for c in cat_filter if c in cat_keys]
            filtered = [c for c in analyzed if c.get("kategorie") in selected_keys]

            with st.spinner("PDF wird erstellt..."):
                path = pdf_export.generate_overview_pdf(filtered)
            with open(path, "rb") as f:
                st.download_button(
                    label="⬇️ Übersicht herunterladen",
                    data=f,
                    file_name="Stormarn_KI_Uebersicht.pdf",
                    mime="application/pdf"
                )

# ──────────────────────────────────────────────────────────
# PAGE: Einstellungen
# ──────────────────────────────────────────────────────────
elif page == "⚙️ Einstellungen":
    st.header("⚙️ Konfiguration")

    st.subheader("Aktuelle config.yaml")
    config_path = "config.yaml"
    if os.path.exists(config_path):
        with open(config_path, "r") as f:
            config_content = f.read()
        st.code(config_content, language="yaml")

        st.info("💡 Um das Thema zu wechseln (z.B. auf 'Wasserstoff'), "
                "einfach die config.yaml bearbeiten und `streamlit run app.py` neu starten.")
    else:
        st.error("config.yaml nicht gefunden.")

    st.markdown("---")
    st.subheader("🔄 Themen-Presets")
    st.markdown("""
Lade eine dieser Beispiel-Konfigurationen um das Radar-Thema zu wechseln:

| Preset | Thema | Region |
|--------|-------|--------|
| `config_ki_stormarn.yaml` | Künstliche Intelligenz | Kreis Stormarn |
| `config_h2_hamburg.yaml` | Wasserstofftechnologie | Hamburg |
| `config_startup.yaml` | Tech-Startups | Berlin |
| `config_nachhaltigkeit.yaml` | Nachhaltigkeit | Schleswig-Holstein |

Einfach umbenennen in `config.yaml` und Streamlit neu starten.
    """)

    st.markdown("---")
    st.subheader("🗑️ Datenbank zurücksetzen")
    if st.checkbox("Ich bin sicher, dass ich alle Daten löschen möchte"):
        if st.button("🗑️ Alle Daten löschen", type="secondary"):
            db_path = "data/radar.db"
            if os.path.exists(db_path):
                os.remove(db_path)
                db.init_db()
                st.success("Datenbank zurückgesetzt.")
            else:
                st.info("Keine Datenbank gefunden.")
